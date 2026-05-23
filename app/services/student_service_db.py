# -*- coding: utf-8 -*-
"""
学生管理服务 - SQLite版本
支持学生名单的增删改查和Excel导入
"""
import os
import csv
import uuid
from datetime import datetime
from app.models.student import Student
from app.models.user import User


def get_all_students():
    """获取所有学生"""
    return Student.get_all()


def get_student_by_id(student_id):
    """根据ID获取学生"""
    return Student.get_by_id(student_id)


def add_student(name, student_no, class_name, **kwargs):
    """添加学生"""
    return Student.create(name, student_no, class_name, **kwargs)


def update_student(student_id, **kwargs):
    """更新学生信息"""
    return Student.update(student_id, **kwargs)


def delete_student(student_id):
    """删除学生（同时删除登录账号）"""
    return Student.delete(student_id)


def batch_delete_students(student_ids):
    """批量删除学生（同时删除登录账号）"""
    return Student.batch_delete(student_ids)


def import_from_csv(file_path):
    """从CSV文件导入学生名单

    支持的列名（不区分大小写）：姓名/名字/name、学号/student_no/student_id、班级/class
    第一行为表头，自动识别列
    """
    imported = 0
    skipped = 0
    errors = []

    try:
        with open(file_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            if not reader.fieldnames:
                return {'imported': 0, 'skipped': 0, 'errors': ['文件为空或格式不正确']}

            # 自动识别列名
            name_col = _find_column(reader.fieldnames, ['姓名', '名字', 'name', '学生姓名'])
            no_col = _find_column(reader.fieldnames, ['学号', 'student_no', 'student_id', 'studentno', '编号'])
            class_col = _find_column(reader.fieldnames, ['班级', 'class', 'class_name', 'classname', '班'])

            if not name_col:
                return {'imported': 0, 'skipped': 0, 'errors': ['未找到"姓名"列，请确保CSV包含姓名列']}

            for i, row in enumerate(reader, start=2):
                try:
                    name = row.get(name_col, '').strip()
                    if not name:
                        skipped += 1
                        continue

                    student_no = row.get(no_col, '').strip() if no_col else ''
                    class_name = row.get(class_col, '').strip() if class_col else ''

                    # 检查学号是否重复
                    if student_no:
                        existing = Student.get_by_student_no(student_no)
                        if existing:
                            skipped += 1
                            continue

                    Student.create(name=name, student_no=student_no, class_name=class_name)
                    imported += 1
                except Exception as e:
                    errors.append(f'第{i}行: {str(e)}')

        return {'imported': imported, 'skipped': skipped, 'errors': errors}

    except UnicodeDecodeError:
        # 尝试 GBK 编码
        try:
            with open(file_path, 'r', encoding='gbk') as f:
                reader = csv.DictReader(f)
                if not reader.fieldnames:
                    return {'imported': 0, 'skipped': 0, 'errors': ['文件为空或格式不正确']}

                name_col = _find_column(reader.fieldnames, ['姓名', '名字', 'name', '学生姓名'])
                no_col = _find_column(reader.fieldnames, ['学号', 'student_no', 'student_id', 'studentno', '编号'])
                class_col = _find_column(reader.fieldnames, ['班级', 'class', 'class_name', 'classname', '班'])

                if not name_col:
                    return {'imported': 0, 'skipped': 0, 'errors': ['未找到"姓名"列']}

                for i, row in enumerate(reader, start=2):
                    try:
                        name = row.get(name_col, '').strip()
                        if not name:
                            skipped += 1
                            continue

                        student_no = row.get(no_col, '').strip() if no_col else ''
                        class_name = row.get(class_col, '').strip() if class_col else ''

                        if student_no:
                            existing = Student.get_by_student_no(student_no)
                            if existing:
                                skipped += 1
                                continue

                        Student.create(name=name, student_no=student_no, class_name=class_name)
                        imported += 1
                    except Exception as e:
                        errors.append(f'第{i}行: {str(e)}')

            return {'imported': imported, 'skipped': skipped, 'errors': errors}
        except Exception as e:
            return {'imported': 0, 'skipped': 0, 'errors': [f'文件读取失败: {str(e)}']}
    except Exception as e:
        return {'imported': 0, 'skipped': 0, 'errors': [f'导入失败: {str(e)}']}


def import_from_excel(file_path):
    """从Excel文件导入学生名单

    支持的列名同CSV
    """
    try:
        import openpyxl
    except ImportError:
        return {'imported': 0, 'skipped': 0, 'errors': ['缺少openpyxl库，请安装: pip install openpyxl']}

    imported = 0
    skipped = 0
    errors = []

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb.active

        # 读取表头
        headers = []
        for cell in ws[1]:
            val = cell.value
            if val:
                headers.append(str(val).strip())
            else:
                headers.append('')

        # 识别列
        name_col = _find_column(headers, ['姓名', '名字', 'name', '学生姓名'])
        no_col = _find_column(headers, ['学号', 'student_no', 'student_id', 'studentno', '编号'])
        class_col = _find_column(headers, ['班级', 'class', 'class_name', 'classname', '班'])

        if name_col is None:
            wb.close()
            return {'imported': 0, 'skipped': 0, 'errors': ['未找到"姓名"列，请确保Excel包含姓名列']}

        name_idx = headers.index(name_col) if name_col else None
        no_idx = headers.index(no_col) if no_col else None
        class_idx = headers.index(class_col) if class_col else None

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                name = str(row[name_idx] or '').strip() if name_idx is not None else ''
                if not name:
                    skipped += 1
                    continue

                student_no = str(row[no_idx] or '').strip() if no_idx is not None else ''
                class_name = str(row[class_idx] or '').strip() if class_idx is not None else ''

                if student_no:
                    existing = Student.get_by_student_no(student_no)
                    if existing:
                        skipped += 1
                        continue

                Student.create(name=name, student_no=student_no, class_name=class_name)
                imported += 1
            except Exception as e:
                errors.append(f'第{i}行: {str(e)}')

        wb.close()
        return {'imported': imported, 'skipped': skipped, 'errors': errors}

    except Exception as e:
        return {'imported': 0, 'skipped': 0, 'errors': [f'导入失败: {str(e)}']}


def _find_column(headers, candidates):
    """在表头中查找匹配的列名"""
    if not headers:
        return None
    headers_lower = [h.lower().strip() for h in headers]
    for candidate in candidates:
        for i, h in enumerate(headers_lower):
            if h == candidate.lower():
                return headers[i]
    return None


def get_class_list():
    """获取所有班级列表"""
    return Student.get_class_list()


def sync_user_accounts():
    """为所有已导入但没有登录账号的学生创建账号"""
    return Student.sync_user_accounts()


def get_students_by_class(class_name):
    """获取指定班级的学生"""
    return Student.get_by_class(class_name)


def search_students(keyword=''):
    """搜索学生（按姓名、学号、班级模糊匹配）"""
    return Student.search(keyword)


def get_student_stats():
    """获取学生统计信息"""
    return Student.get_stats()
