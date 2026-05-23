# -*- coding: utf-8 -*-
"""
学生管理服务
支持学生名单的增删改查和Excel导入
"""
import os
import json
import csv
import uuid
from datetime import datetime

# 数据存储路径
DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), 'data')
STUDENTS_FILE = os.path.join(DATA_DIR, 'students.json')


def _load_students():
    """加载学生数据"""
    if os.path.exists(STUDENTS_FILE):
        with open(STUDENTS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return []


def _save_students(students):
    """保存学生数据"""
    os.makedirs(DATA_DIR, exist_ok=True)
    with open(STUDENTS_FILE, 'w', encoding='utf-8') as f:
        json.dump(students, f, ensure_ascii=False, indent=2)


def get_all_students():
    """获取所有学生"""
    return _load_students()


def get_student_by_id(student_id):
    """根据ID获取学生"""
    students = _load_students()
    for s in students:
        if s.get('id') == student_id:
            return s
    return None


def add_student(name, student_no, class_name, **kwargs):
    """添加学生"""
    students = _load_students()
    student = {
        'id': str(uuid.uuid4())[:8],
        'name': name.strip(),
        'student_no': student_no.strip(),
        'class_name': class_name.strip(),
        'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    }
    # 额外字段
    for key in ('gender', 'phone', 'email', 'remark'):
        if key in kwargs and kwargs[key]:
            student[key] = str(kwargs[key]).strip()
    students.append(student)
    _save_students(students)
    return student


def update_student(student_id, **kwargs):
    """更新学生信息"""
    students = _load_students()
    for s in students:
        if s.get('id') == student_id:
            for key in ('name', 'student_no', 'class_name', 'gender', 'phone', 'email', 'remark'):
                if key in kwargs:
                    s[key] = str(kwargs[key]).strip()
            s['updated_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            _save_students(students)
            return s
    return None


def delete_student(student_id):
    """删除学生（同时删除登录账号）"""
    students = _load_students()
    deleted_student = None
    new_students = []
    for s in students:
        if s.get('id') == student_id:
            deleted_student = s
        else:
            new_students.append(s)
    if deleted_student:
        _save_students(new_students)
        # 同步删除登录账号
        student_no = deleted_student.get('student_no', '')
        if student_no:
            _delete_user_account(student_no)
        return True
    return False


def batch_delete_students(student_ids):
    """批量删除学生（同时删除登录账号）"""
    students = _load_students()
    id_set = set(student_ids)
    new_students = []
    deleted_nos = []
    for s in students:
        if s.get('id') in id_set:
            no = s.get('student_no', '')
            if no:
                deleted_nos.append(no)
        else:
            new_students.append(s)
    deleted = len(students) - len(new_students)
    _save_students(new_students)
    # 同步删除登录账号
    for no in deleted_nos:
        _delete_user_account(no)
    return deleted


def import_from_csv(file_path):
    """从CSV文件导入学生名单

    支持的列名（不区分大小写）：姓名/名字/name、学号/student_no/student_id、班级/class
    第一行为表头，自动识别列
    """
    students = _load_students()
    imported = 0
    skipped = 0
    errors = []

    try:
        with open(file_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            if not reader.fieldnames:
                return {'imported': 0, 'skipped': 0, 'errors': ['文件为空或格式不正确']}

            # 自动识别列名
            name_col = _find_column(reader.fieldnames, ['姓名', '名字', 'name', '学生姓名'])
            no_col = _find_column(reader.fieldnames, ['学号', 'student_no', 'student_id', 'studentno', '编号'])
            class_col = _find_column(reader.fieldnames, ['班级', 'class', 'class_name', 'classname', '班'])

            if not name_col:
                return {'imported': 0, 'skipped': 0, 'errors': ['未找到"姓名"列，请确保CSV包含姓名列']}

            for i, row in enumerate(reader, start=2):
                try:
                    name = row.get(name_col, '').strip()
                    if not name:
                        skipped += 1
                        continue

                    student_no = row.get(no_col, '').strip() if no_col else ''
                    class_name = row.get(class_col, '').strip() if class_col else ''

                    # 检查学号是否重复
                    if student_no:
                        exists = any(s.get('student_no') == student_no for s in students)
                        if exists:
                            skipped += 1
                            continue

                    student = {
                        'id': str(uuid.uuid4())[:8],
                        'name': name,
                        'student_no': student_no,
                        'class_name': class_name,
                        'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    }
                    students.append(student)

                    # 自动创建登录账号（用户名=学号，密码=学号后6位）
                    if student_no:
                        _create_user_account(student_no, name)

                    imported += 1
                except Exception as e:
                    errors.append(f'第{i}行: {str(e)}')

        _save_students(students)
        return {'imported': imported, 'skipped': skipped, 'errors': errors}

    except UnicodeDecodeError:
        # 尝试 GBK 编码
        try:
            with open(file_path, 'r', encoding='gbk') as f:
                reader = csv.DictReader(f)
                if not reader.fieldnames:
                    return {'imported': 0, 'skipped': 0, 'errors': ['文件为空或格式不正确']}

                name_col = _find_column(reader.fieldnames, ['姓名', '名字', 'name', '学生姓名'])
                no_col = _find_column(reader.fieldnames, ['学号', 'student_no', 'student_id', 'studentno', '编号'])
                class_col = _find_column(reader.fieldnames, ['班级', 'class', 'class_name', 'classname', '班'])

                if not name_col:
                    return {'imported': 0, 'skipped': 0, 'errors': ['未找到"姓名"列']}

                for i, row in enumerate(reader, start=2):
                    try:
                        name = row.get(name_col, '').strip()
                        if not name:
                            skipped += 1
                            continue

                        student_no = row.get(no_col, '').strip() if no_col else ''
                        class_name = row.get(class_col, '').strip() if class_col else ''

                        if student_no:
                            exists = any(s.get('student_no') == student_no for s in students)
                            if exists:
                                skipped += 1
                                continue

                        student = {
                            'id': str(uuid.uuid4())[:8],
                            'name': name,
                            'student_no': student_no,
                            'class_name': class_name,
                            'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        }
                        students.append(student)
                        imported += 1
                    except Exception as e:
                        errors.append(f'第{i}行: {str(e)}')

            _save_students(students)
            return {'imported': imported, 'skipped': skipped, 'errors': errors}
        except Exception as e:
            return {'imported': 0, 'skipped': 0, 'errors': [f'文件读取失败: {str(e)}']}
    except Exception as e:
        return {'imported': 0, 'skipped': 0, 'errors': [f'导入失败: {str(e)}']}


def import_from_excel(file_path):
    """从Excel文件导入学生名单

    支持的列名同CSV
    """
    try:
        import openpyxl
    except ImportError:
        return {'imported': 0, 'skipped': 0, 'errors': ['缺少openpyxl库，请安装: pip install openpyxl']}

    students = _load_students()
    imported = 0
    skipped = 0
    errors = []

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb.active

        # 读取表头
        headers = []
        for cell in ws[1]:
            val = cell.value
            if val:
                headers.append(str(val).strip())
            else:
                headers.append('')

        # 识别列
        name_col = _find_column(headers, ['姓名', '名字', 'name', '学生姓名'])
        no_col = _find_column(headers, ['学号', 'student_no', 'student_id', 'studentno', '编号'])
        class_col = _find_column(headers, ['班级', 'class', 'class_name', 'classname', '班'])

        if name_col is None:
            wb.close()
            return {'imported': 0, 'skipped': 0, 'errors': ['未找到"姓名"列，请确保Excel包含姓名列']}

        name_idx = headers.index(name_col) if name_col else None
        no_idx = headers.index(no_col) if no_col else None
        class_idx = headers.index(class_col) if class_col else None

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                name = str(row[name_idx] or '').strip() if name_idx is not None else ''
                if not name:
                    skipped += 1
                    continue

                student_no = str(row[no_idx] or '').strip() if no_idx is not None else ''
                class_name = str(row[class_idx] or '').strip() if class_idx is not None else ''

                if student_no:
                    exists = any(s.get('student_no') == student_no for s in students)
                    if exists:
                        skipped += 1
                        continue

                student = {
                    'id': str(uuid.uuid4())[:8],
                    'name': name,
                    'student_no': student_no,
                    'class_name': class_name,
                    'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                }
                students.append(student)

                # 自动创建登录账号（用户名=学号，密码=学号后6位）
                if student_no:
                    _create_user_account(student_no, name)

                imported += 1
            except Exception as e:
                errors.append(f'第{i}行: {str(e)}')

        wb.close()
        _save_students(students)
        return {'imported': imported, 'skipped': skipped, 'errors': errors}

    except Exception as e:
        return {'imported': 0, 'skipped': 0, 'errors': [f'导入失败: {str(e)}']}


def _find_column(headers, candidates):
    """在表头中查找匹配的列名"""
    if not headers:
        return None
    headers_lower = [h.lower().strip() for h in headers]
    for candidate in candidates:
        for i, h in enumerate(headers_lower):
            if h == candidate.lower():
                return headers[i]
    return None


def _create_user_account(student_no, name):
    """自动创建登录账号（用户名=学号，密码=学号后6位）"""
    try:
        import hashlib
        # 使用与auth.py相同的路径：app/data/users.json
        users_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'data')
        users_file = os.path.join(users_dir, 'users.json')
        users = []
        if os.path.exists(users_file):
            with open(users_file, 'r', encoding='utf-8') as f:
                users = json.load(f)

        # 检查是否已存在
        if student_no in users:
            return

        password = student_no[-6:] if len(student_no) >= 6 else student_no
        password_hash = hashlib.sha256(password.encode('utf-8')).hexdigest()
        user = {
            'password_hash': password_hash,
            'name': name,
            'role': 'student',
            'title': '学生',
        }
        users[student_no] = user
        os.makedirs(users_dir, exist_ok=True)
        with open(users_file, 'w', encoding='utf-8') as f:
            json.dump(users, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def _delete_user_account(student_no):
    """删除登录账号"""
    try:
        users_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'data')
        users_file = os.path.join(users_dir, 'users.json')
        if not os.path.exists(users_file):
            return
        with open(users_file, 'r', encoding='utf-8') as f:
            users = json.load(f)
        if student_no in users:
            del users[student_no]
            with open(users_file, 'w', encoding='utf-8') as f:
                json.dump(users, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def get_class_list():
    """获取所有班级列表"""
    students = _load_students()
    classes = sorted(set(s.get('class_name', '') for s in students if s.get('class_name')))
    return classes


def sync_user_accounts():
    """为所有已导入但没有登录账号的学生创建账号"""
    students = _load_students()
    created = 0
    for s in students:
        student_no = s.get('student_no', '')
        name = s.get('name', '')
        if student_no and name:
            _create_user_account(student_no, name)
            created += 1
    return created


def get_students_by_class(class_name):
    """获取指定班级的学生"""
    students = _load_students()
    return [s for s in students if s.get('class_name') == class_name]


def search_students(keyword=''):
    """搜索学生（按姓名、学号、班级模糊匹配）"""
    students = _load_students()
    if not keyword:
        return students
    keyword = keyword.lower()
    return [
        s for s in students
        if keyword in s.get('name', '').lower()
        or keyword in s.get('student_no', '').lower()
        or keyword in s.get('class_name', '').lower()
    ]


def get_student_stats():
    """获取学生统计信息"""
    students = _load_students()
    classes = get_class_list()
    return {
        'total': len(students),
        'class_count': len(classes),
        'classes': classes,
        'class_distribution': {
            c: len([s for s in students if s.get('class_name') == c])
            for c in classes
        }
    }
